import os
from datetime import datetime
from utils.date_utils import get_now_lima
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

# Directorio de exportaciones absoluto
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
EXPORTS_DIR = os.path.join(BASE_DIR, "exports")

_COLOR_HEADER     = "1a1a2e"
_COLOR_FILA_PAR   = "f5f5f5"
_COLOR_FILA_IMPAR = "ffffff"
_COLOR_ACENTO     = "e8e8f0"

def _borde_fino():
    lado = Side(style="thin", color="cccccc")
    return Border(left=lado, right=lado, top=lado, bottom=lado)

def _borde_medio():
    lado = Side(style="medium", color="1a1a2e")
    return Border(left=lado, right=lado, top=lado, bottom=lado)


def generar_excel_sesion(activos: list, sesion: dict = None) -> str:
    """
    Genera el Excel de resumen (de sesión o global) y lo guarda en exports/.
    Retorna la ruta del archivo generado.
    """
    os.makedirs(EXPORTS_DIR, exist_ok=True)

    now = get_now_lima()
    timestamp = now.strftime("%Y%m%d_%H%M%S")
    prefijo = f"sesion_{sesion.get('id', 'x')}" if sesion else "inventario_global"
    nombre_archivo = f"{prefijo}_{timestamp}.xlsx"
    ruta = os.path.join(EXPORTS_DIR, nombre_archivo)

    wb = Workbook()

    _hoja_resumen(wb, activos, sesion)
    _hoja_activos(wb, activos)

    wb.save(ruta)
    return ruta


def _hoja_resumen(wb: Workbook, activos: list, sesion: dict):
    ws = wb.active
    ws.title = "Resumen"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 35

    fill_header = PatternFill("solid", fgColor=_COLOR_HEADER)
    fill_acento  = PatternFill("solid", fgColor=_COLOR_ACENTO)
    font_blanco  = Font(name="Calibri", bold=True, color="ffffff", size=13)
    font_titulo  = Font(name="Calibri", bold=True, color=_COLOR_HEADER, size=11)
    font_valor   = Font(name="Calibri", color="333333", size=10)
    alineado_c   = Alignment(horizontal="center", vertical="center")
    alineado_i   = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # Título
    ws.merge_cells("A1:B1")
    ws["A1"] = "Tecsup — Departamento de Tecnología Digital"
    ws["A1"].font = font_blanco
    ws["A1"].fill = fill_header
    ws["A1"].alignment = alineado_c
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:B2")
    ws["A2"] = f"Resumen de {'Sesión de Inventariado' if sesion else 'Inventario Global'}"
    ws["A2"].font = Font(name="Calibri", color="888888", size=10, italic=True)
    ws["A2"].fill = fill_header
    ws["A2"].alignment = alineado_c
    ws.row_dimensions[2].height = 18

    # Separador
    ws.row_dimensions[3].height = 8

    if sesion:
        ubicacion = " / ".join(filter(None, [
            sesion.get("pabellon"),
            sesion.get("laboratorio"),
            sesion.get("armario"),
        ])) or "No especificada"
        fecha = (sesion.get("creada_en") or "")[:10]
        alcance_label = "Ubicación"
    else:
        ubicacion = "Inventario Completo (Todo el campus)"
        fecha = get_now_lima().strftime("%Y-%m-%d")
        alcance_label = "Alcance"

    datos = [
        ("Técnico Responsable", sesion.get("tecnico") or "N/A") if sesion else ("Generado por", "Sistema Inventario"),
        (alcance_label, ubicacion),
        ("Fecha de corte", fecha),
        ("Total de activos registrados", str(len(activos))),
        ("Registrados por OCR", str(sum(1 for a in activos if a.get("origen") == "ocr"))),
        ("Registrados por voz", str(sum(1 for a in activos if a.get("origen") == "voz"))),
        ("Registrados manualmente", str(sum(1 for a in activos if a.get("origen") == "manual"))),
        ("Documento generado", get_now_lima().strftime("%d/%m/%Y %H:%M:%S")),
    ]

    for i, (label, valor) in enumerate(datos, start=4):
        ws.row_dimensions[i].height = 22
        celd_label = ws.cell(row=i, column=1, value=label)
        celd_valor = ws.cell(row=i, column=2, value=valor)

        celd_label.font = font_titulo
        celd_valor.font = font_valor
        celd_label.alignment = alineado_i
        celd_valor.alignment = alineado_i

        if i % 2 == 0:
            celd_label.fill = fill_acento
            celd_valor.fill = fill_acento

        celd_label.border = _borde_fino()
        celd_valor.border = _borde_fino()


def _hoja_activos(wb: Workbook, activos: list):
    ws = wb.create_sheet(title="Activos")
    ws.sheet_view.showGridLines = False

    columnas = [
        ("#",           5),
        ("Nombre",      28),
        ("Marca",       14),
        ("Modelo",      18),
        ("Tipo",        16),
        ("N° Serie",    18),
        ("Estado",      12),
        ("Ubicación",   22),
        ("Técnico",     20),
        ("Origen",      10),
        ("Registrado",  18),
    ]

    for col_idx, (titulo, ancho) in enumerate(columnas, start=1):
        letra = get_column_letter(col_idx)
        ws.column_dimensions[letra].width = ancho

        celda = ws.cell(row=1, column=col_idx, value=titulo)
        celda.font       = Font(name="Calibri", bold=True, color="ffffff", size=10)
        celda.fill       = PatternFill("solid", fgColor=_COLOR_HEADER)
        celda.alignment  = Alignment(horizontal="center", vertical="center")
        celda.border     = _borde_medio()

    ws.row_dimensions[1].height = 22

    fill_par   = PatternFill("solid", fgColor=_COLOR_FILA_PAR)
    fill_impar = PatternFill("solid", fgColor=_COLOR_FILA_IMPAR)
    font_normal = Font(name="Calibri", size=9)
    font_serial = Font(name="Calibri", size=9, bold=True)

    for fila_idx, activo in enumerate(activos, start=2):
        fill = fill_par if fila_idx % 2 == 0 else fill_impar
        ws.row_dimensions[fila_idx].height = 18

        valores = [
            fila_idx - 1,
            activo.get("nombre") or "",
            activo.get("marca") or "",
            activo.get("modelo") or "",
            activo.get("tipo") or "",
            activo.get("numero_serie") or "",
            activo.get("estado") or "",
            activo.get("ubicacion") or "",
            activo.get("tecnico") or "",
            activo.get("origen") or "",
            (activo.get("creado_en") or "")[:16],
        ]

        for col_idx, valor in enumerate(valores, start=1):
            celda = ws.cell(row=fila_idx, column=col_idx, value=valor)
            # El número de serie va en bold para destacarlo
            celda.font = font_serial if col_idx == 6 else font_normal
            celda.fill = fill
            celda.alignment = Alignment(
                horizontal="center" if col_idx in (1, 7, 8, 9) else "left",
                vertical="center",
                wrap_text=True,
            )
            celda.border = _borde_fino()

    # Autofilter en la fila de encabezados
    ws.auto_filter.ref = f"A1:{get_column_letter(len(columnas))}1"